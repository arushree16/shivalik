from flask import Flask, render_template, request, redirect, url_for, flash, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
import openpyxl
import random
from openpyxl.styles import Alignment, Border, Side, Font
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, IntegerField, SubmitField
from wtforms.validators import DataRequired, Length, NumberRange, EqualTo, ValidationError
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from PyPDF2 import PdfMerger
from xhtml2pdf import pisa
import tempfile
import zipfile

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'instance', 'school.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

if not os.path.exists(os.path.join(basedir, 'static', 'templates')):
    os.makedirs(os.path.join(basedir, 'static', 'templates'))

class Teacher(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), unique=True, nullable=False)
    password = db.Column(db.String(60), nullable=False)
    name = db.Column(db.String(100), nullable=True)  
    assigned_class = db.Column(db.String(10), nullable=False)

    def __repr__(self):
        return f"Teacher('{self.username}', '{self.assigned_class}')"

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    gender = db.Column(db.String(10))
    category = db.Column(db.String(20))
    religion = db.Column(db.String(50))
    fathers_name = db.Column(db.String(100))
    class_name = db.Column(db.String(20))
    roll_no = db.Column(db.String(20))
    admission_no = db.Column(db.String(20))
    attendance = db.Column(db.Integer, default=80)  # Default attendance is 80%
    
    # Academic subjects
    english_pa = db.Column(db.Integer, default=0)
    english_ma = db.Column(db.Integer, default=0)
    english_portfolio = db.Column(db.Integer, default=0)
    english_se = db.Column(db.Integer, default=0)
    english_term = db.Column(db.Integer, default=0)
    
    hindi_pa = db.Column(db.Integer, default=0)
    hindi_ma = db.Column(db.Integer, default=0)
    hindi_portfolio = db.Column(db.Integer, default=0)
    hindi_se = db.Column(db.Integer, default=0)
    hindi_term = db.Column(db.Integer, default=0)
    
    maths_pa = db.Column(db.Integer, default=0)
    maths_ma = db.Column(db.Integer, default=0)
    maths_portfolio = db.Column(db.Integer, default=0)
    maths_se = db.Column(db.Integer, default=0)
    maths_term = db.Column(db.Integer, default=0)
    
    science_pa = db.Column(db.Integer, default=0)
    science_ma = db.Column(db.Integer, default=0)
    science_portfolio = db.Column(db.Integer, default=0)
    science_se = db.Column(db.Integer, default=0)
    science_term = db.Column(db.Integer, default=0)
    
    sst_pa = db.Column(db.Integer, default=0)
    sst_ma = db.Column(db.Integer, default=0)
    sst_portfolio = db.Column(db.Integer, default=0)
    sst_se = db.Column(db.Integer, default=0)
    sst_term = db.Column(db.Integer, default=0)
    
    # Co-curricular subject with marks
    cocurricular_pa = db.Column(db.Integer, default=0)
    cocurricular_ma = db.Column(db.Integer, default=0)
    cocurricular_portfolio = db.Column(db.Integer, default=0)
    cocurricular_se = db.Column(db.Integer, default=0)
    cocurricular_term = db.Column(db.Integer, default=0)
    
    # Activity grades
    gk_grade = db.Column(db.String(2))
    computer_grade = db.Column(db.String(2))
    drawing_grade = db.Column(db.String(2))

class RegistrationForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=2, max=20)])
    password = PasswordField('Password', validators=[DataRequired()])
    confirm_password = PasswordField('Confirm Password', validators=[DataRequired(), EqualTo('password')])
    name = StringField('Name', validators=[Length(max=100)])  # Optional name field
    assigned_class = StringField('Assigned Class', validators=[DataRequired()])
    submit = SubmitField('Sign Up')

    def validate_username(self, username):
        teacher = Teacher.query.filter_by(username=username.data).first()
        if teacher:
            raise ValidationError('That username is taken. Please choose a different one.')

class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired(), Length(min=2, max=20)])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Login')

@login_manager.user_loader
def load_user(user_id):
    return Teacher.query.get(int(user_id))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = RegistrationForm()
    print("\nRegistration Debug Info:")
    print(f"Form submitted: {request.method == 'POST'}")
    print(f"Form validated: {form.validate_on_submit()}")
    
    if not form.validate():
        print("Form validation errors:", form.errors)
    
    if form.validate_on_submit():
        try:
            username = form.username.data.lower()
            print(f"Processing registration for username: {username}")
            
            hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
            teacher = Teacher(
                username=username,
                password=hashed_password,
                name=form.name.data if form.name.data else None,
                assigned_class=form.assigned_class.data
            )
            
            print("Teacher object created, attempting to save to database...")
            db.session.add(teacher)
            db.session.commit()
            
            print(f"Successfully registered user: {teacher.username}")
            flash('Registered successfully! Please login.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            print(f"Registration error: {str(e)}")
            db.session.rollback()
            flash('Registration failed. Please try again.', 'danger')
    else:
        print("Form validation failed or not submitted")
    
    return render_template('register.html', title='Register', form=form)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data.lower()
        print("\nLogin Debug Info:")
        print(f"Attempting login with username: {username}")
        
        # Print all teachers in database
        all_teachers = Teacher.query.all()
        print("\nAll registered teachers:")
        for t in all_teachers:
            print(f"ID: {t.id}, Username: {t.username}, Name: {t.name}")
        
        teacher = Teacher.query.filter_by(username=username).first()
        
        if teacher:
            print(f"\nFound matching teacher: {teacher.username}")
            if bcrypt.check_password_hash(teacher.password, form.password.data):
                login_user(teacher)
                flash('Login successful!', 'success')
                next_page = request.args.get('next')
                return redirect(next_page) if next_page else redirect(url_for('index'))
            else:
                print("Password verification failed")
                flash('Login unsuccessful. Incorrect password.', 'danger')
        else:
            print(f"No teacher found with username: {username}")
            flash('Login unsuccessful. Username not found.', 'danger')
    
    return render_template('login.html', title='Login', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully.', 'success')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html', assigned_class=current_user.assigned_class)

@app.route('/add_student', methods=['GET', 'POST'])
@login_required
def add_student():
    if not current_user.is_authenticated:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('login'))

    if request.method == 'POST':
        # Get form data
        name = request.form.get('name')
        gender = request.form.get('gender')
        category = request.form.get('category')
        religion = request.form.get('religion')
        fathers_name = request.form.get('fathers_name')
        roll_no = request.form.get('roll_no')
        admission_no = request.form.get('admission_no')
        class_name = current_user.assigned_class

        # Create new student
        # Get attendance from form, default to 80 if not provided
        attendance = request.form.get('attendance', 80, type=int)
        # Ensure attendance is between 80 and 100
        attendance = min(max(80, attendance), 100)

        student = Student(
            name=name,
            gender=gender,
            category=category,
            religion=religion,
            fathers_name=fathers_name,
            roll_no=roll_no,
            admission_no=admission_no,
            class_name=class_name,
            attendance=attendance
        )

        # Max marks for each component
        max_marks = {
            'pa': 10,
            'ma': 10,
            'portfolio': 10,
            'se': 10,  # SE is out of 10
            'term': 60
        }

        # Add academic marks for all subjects
        subjects = ['english', 'hindi', 'maths', 'science', 'sst', 'cocurricular']
        components = ['pa', 'ma', 'portfolio', 'se', 'term']

        for subject in subjects:
            for component in components:
                # Get mark from form
                mark = request.form.get(f'{subject}_{component}', 0, type=int)
                # Ensure mark doesn't exceed max
                mark = min(max(0, mark), max_marks[component])
                # Set the mark
                setattr(student, f'{subject}_{component}', mark)

        # Add activity grades
        student.gk_grade = request.form.get("gk_grade", "")
        student.computer_grade = request.form.get("computer_grade", "")
        student.drawing_grade = request.form.get("drawing_grade", "")

        db.session.add(student)
        db.session.commit()

        flash('Student added successfully!', 'success')
        return redirect(url_for('view_class', class_name=class_name))

    return render_template('add_student.html')

@app.route('/view_class/<class_name>')
@login_required
def view_class(class_name):
    if class_name != current_user.assigned_class:
        flash('You can only view your assigned class!', 'error')
        return redirect(url_for('index'))
    
    students = Student.query.filter_by(class_name=class_name).all()
    return render_template('view_class.html', students=students, class_name=class_name)

@app.route('/view_marksheet/<class_name>')
@login_required
def view_marksheet(class_name):
    if class_name != current_user.assigned_class:
        flash('You can only view marksheets of your assigned class!', 'error')
        return redirect(url_for('index'))
    
    students = Student.query.filter_by(class_name=class_name).all()
    for student in students:
        # Calculate totals for each subject
        student_total = 0
        for subject in ['english', 'hindi', 'maths', 'science', 'sst']:
            # Get component marks
            pa = getattr(student, f"{subject}_pa", 0) or 0
            ma = getattr(student, f"{subject}_ma", 0) or 0
            portfolio = getattr(student, f"{subject}_portfolio", 0) or 0
            se = getattr(student, f"{subject}_se", 0) or 0
            term = getattr(student, f"{subject}_term", 0) or 0
            
            # Calculate subject total (PA:10 + MA:10 + Port:10 + SE:10 + Term:60)
            component_total = pa + ma + portfolio + se  # Sum of 10-mark components
            subject_total = component_total + term  # Add term marks (out of 60)
            
            # Store subject total
            setattr(student, f"{subject}_total", subject_total)
            student_total += subject_total
        
        # Set the total and calculate percentage
        student.total = student_total
        student.percentage = round((student_total / 500) * 100, 2)  # 500 = 5 subjects * 100 marks
        
        # Calculate grade based on percentage
        if student.percentage >= 91:
            student.grade = 'A+'
        elif student.percentage >= 81:
            student.grade = 'A'
        elif student.percentage >= 71:
            student.grade = 'B+'
        elif student.percentage >= 61:
            student.grade = 'B'
        elif student.percentage >= 51:
            student.grade = 'C+'
        elif student.percentage >= 41:
            student.grade = 'C'
        else:
            student.grade = 'D'
        
        student.passed = student.percentage >= 33  # Pass criteria is 33%

    total_students = len(students)
    promoted_students = len([s for s in students if s.passed])
    result_percentage = round((promoted_students / total_students * 100), 2) if total_students > 0 else 0

    return render_template('marksheet.html', 
                         students=students,
                         teacher=current_user,
                         class_name=class_name,
                         total_students=total_students,
                         promoted_students=promoted_students,
                         result_percentage=result_percentage)

@app.route('/view_report_card/<int:student_id>')
@login_required
def view_report_card(student_id):
    student = Student.query.get_or_404(student_id)
    
    if student.class_name != current_user.assigned_class:
        flash('You can only view report cards for students in your assigned class!', 'error')
        return redirect(url_for('index'))
    
    try:
        # Load the template Excel file
        template_path = os.path.join(basedir, 'static', 'templates', 'report_template.xlsx')
        wb = openpyxl.load_workbook(template_path, data_only=False)  # Keep formulas
        ws = wb.active
        
        def safe_write_to_cell(ws, cell_coord, value):
            try:
                cell = ws[cell_coord]
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Find the master cell of the merged range
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            master_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            master_cell.value = value
                            return
                else:
                    cell.value = value
            except Exception as e:
                print(f"Error writing to cell {cell_coord}: {str(e)}")
                # Try writing directly using row/column
                row = openpyxl.utils.cell.coordinate_to_tuple(cell_coord)[0]
                col = openpyxl.utils.cell.coordinate_to_tuple(cell_coord)[1]
                ws.cell(row=row, column=col, value=value)
        
        # Fill student details
        safe_write_to_cell(ws, 'C5', student.name)  # Name in column C
        safe_write_to_cell(ws, 'E5', student.class_name)  # Class in column E
        safe_write_to_cell(ws, 'G5', student.roll_no)  # Roll No in column G
        safe_write_to_cell(ws, 'I5', student.admission_no)  # Admission No in column I
        safe_write_to_cell(ws, 'C6', student.fathers_name)  # Father's Name in column C
        
        # Subject rows are already defined in template
        subjects = [
            ('ENGLISH', 10, 'english'),
            ('HINDI', 11, 'hindi'),
            ('MATHS', 12, 'maths'),
            ('SCIENCE', 13, 'science'),
            ('SOCIAL SCIENCE', 14, 'sst'),
            ('CO-CURRICULAR', 15, 'cocurricular')  # Added co-curricular subject
        ]
        
        # Components and their columns
        components = [
            ('pa', 'B'),      # PA in column B
            ('ma', 'C'),      # MA in column C
            ('portfolio', 'D'), # Portfolio in column D
            ('se', 'E'),      # SE in column E
            ('term', 'F')     # Term in column F
        ]
        
        # Fill marks for each subject
        for subject_name, row, code in subjects:
            for component, col in components:
                mark = getattr(student, f"{code}_{component}", 0) or 0
                safe_write_to_cell(ws, f"{col}{row}", mark)
        
        # Fill co-scholastic grades in exact cells
        safe_write_to_cell(ws, 'C17', student.gk_grade or 'C')
        safe_write_to_cell(ws, 'F17', student.computer_grade or 'C')
        safe_write_to_cell(ws, 'I17', student.drawing_grade or 'C')
        
        # Save to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            return send_file(
                tmp_path,
                as_attachment=True,
                download_name=f"report_card_{student.name}_{student.class_name}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            print(f"Error sending file: {str(e)}")
            flash('Error generating report card. Please try again.', 'error')
            return redirect(url_for('view_class', class_name=student.class_name))
        finally:
            # Clean up
            wb.close()
            try:
                os.unlink(tmp_path)
            except:
                pass
            
    except Exception as e:
        print(f"Error generating report card: {str(e)}")
        flash('Error generating report card. Please try again.', 'error')
        return redirect(url_for('view_class', class_name=student.class_name))

@app.route('/download_marks_list/<class_name>')
@login_required
def download_marks_list(class_name):
    if class_name != current_user.assigned_class:
        flash('You can only download marks list for your assigned class!', 'error')
        return redirect(url_for('index'))
    
    students = Student.query.filter_by(class_name=class_name).all()
    
    # Load the template Excel file
    template_path = os.path.join(basedir, 'static', 'templates', 'marksheet_template.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # Helper function to safely set cell values for both regular and merged cells
    def set_cell_value(row, col, value):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            # Find the master cell of the merged region
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    master_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    master_cell.value = value
                    return
        else:
            cell.value = value
    
    # Update class and teacher info
    set_cell_value(4, 1, f"CLASS TEACHER NAME: {current_user.name}")
    set_cell_value(4, 4, f"CLASS: {class_name}")
    
    # Start filling student data from row 8 (after headers)
    start_row = 8
    for idx, student in enumerate(students, 1):
        row = start_row + idx - 1
        
        # Basic info
        ws.cell(row=row, column=1, value=idx)  # S.No.
        ws.cell(row=row, column=2, value=student.admission_no)
        ws.cell(row=row, column=3, value=student.name)
        ws.cell(row=row, column=4, value=student.fathers_name)
        ws.cell(row=row, column=5, value=student.category)
        ws.cell(row=row, column=6, value=student.gender)
        ws.cell(row=row, column=7, value=student.religion)
        ws.cell(row=row, column=8, value=f"{student.attendance}%")  # Column H for attendance
        
        # Map subjects to their Excel columns (matching template exactly)
        # Each subject has 5 columns: PA(10), MA(10), Portfolio(10), SE(10), Term(60)
        # Column mapping based on Excel formulas:
        # Hindi: I-M, total in N, grade in O
        # English: P-T, total in U, grade in V
        # Maths: W-AA, total in AB, grade in AC
        # EVS: AD-AH, total in AI, grade in AJ
        # SST: AK-AO, total in AP, grade in AQ
        subject_columns = [
            ('hindi', 'I'),      # Hindi starts at I
            ('english', 'P'),    # English starts at P
            ('maths', 'W'),      # Maths starts at W
            ('science', 'AD'),   # Science (EVS) starts at AD
            ('sst', 'AK'),      # SST starts at AK
            ('cocurricular', 'AR')  # Co-curricular starts at AR
        ]

        # Components for each subject and their max marks
        components = [
            ('pa', 10),
            ('ma', 10),
            ('portfolio', 10),
            ('se', 10),  # SE is out of 10
            ('term', 60)
        ]

        # Fill in marks for each subject
        for subject, start_col in subject_columns:
            # Convert column letter to number (e.g., 'AD' -> 30)
            if len(start_col) == 1:
                col = ord(start_col) - ord('A') + 1
            else:
                col = (ord(start_col[0]) - ord('A') + 1) * 26 + (ord(start_col[1]) - ord('A') + 1)
            
            # Fill in each component's marks
            for i, (component, max_marks) in enumerate(components):
                # Get mark from student object
                mark = getattr(student, f"{subject}_{component}", 0)
                if mark is not None:
                    # Ensure mark doesn't exceed max
                    mark = min(mark, max_marks)
                    ws.cell(row=row, column=col + i, value=mark)
        

    
    # Update summary at bottom using set_cell_value function
    summary_row = start_row + len(students) + 2
    
    # Calculate promoted students count
    promoted_count = len([s for s in students if sum([
        getattr(s, f"{subj}_{comp}", 0) or 0
        for subj in ['hindi', 'english', 'maths', 'science', 'sst']
        for comp in ['pa', 'ma', 'portfolio', 'se', 'term']
    ]) >= 200])
    
    # Calculate result percentage
    result_percentage = (promoted_count/len(students)*100) if students else 0
    
    # Set summary information
    set_cell_value(summary_row, 1, f"TOTAL STUDENTS: {len(students)}")
    set_cell_value(summary_row, 4, f"TOTAL STUDENTS PROMOTED: {promoted_count}")
    set_cell_value(summary_row, 7, f"RESULT PERCENTAGE: {result_percentage:.1f}%")
    
    # Save the file
    filename = f"Class_{class_name}_Marks_List.xlsx"
    wb.save(filename)
    
    # Just return the file without redirecting
    return send_file(
        filename,
        as_attachment=True,
        download_name=f"Class_{class_name}_Marks_List.xlsx"
    )

@app.route('/edit_student/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    # Check if teacher has permission to edit this student
    if student.class_name != current_user.assigned_class:
        flash('You can only edit students from your assigned class!', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        student.name = request.form['name']
        student.gender = request.form['gender']
        student.category = request.form['category']
        student.religion = request.form['religion']
        student.fathers_name = request.form['fathers_name']
        student.roll_no = request.form['roll_no']
        student.admission_no = request.form['admission_no']
        
        # Update attendance (between 80 and 100)
        attendance = request.form.get('attendance', 80, type=int)
        student.attendance = min(max(80, attendance), 100)
        
        # Max marks for each component
        max_marks = {
            'pa': 10,
            'ma': 10,
            'portfolio': 10,
            'se': 10,  # SE is out of 10
            'term': 60
        }

        # Update academic marks for all subjects
        subjects = ['english', 'hindi', 'maths', 'science', 'sst', 'cocurricular']
        components = ['pa', 'ma', 'portfolio', 'se', 'term']

        for subject in subjects:
            for component in components:
                # Get mark from form
                mark = request.form.get(f'{subject}_{component}', 0, type=int)
                # Ensure mark doesn't exceed max
                mark = min(max(0, mark), max_marks[component])
                # Set the mark
                setattr(student, f'{subject}_{component}', mark)
        
        student.gk_grade = request.form.get('gk_grade', 'E')
        student.computer_grade = request.form.get('computer_grade', 'E')
        student.drawing_grade = request.form.get('drawing_grade', 'E')
        
        db.session.commit()
        flash('Student information updated successfully!', 'success')
        return redirect(url_for('view_class', class_name=student.class_name))
    
    return render_template('edit_student.html', student=student)

@app.route('/delete_student/<int:student_id>', methods=['GET', 'POST'])
@login_required
def delete_student(student_id):
    student = Student.query.get_or_404(student_id)
    
    # Check if the student belongs to the teacher's class
    if str(student.class_name) != str(current_user.assigned_class):
        flash('You can only delete students from your assigned class!', 'error')
        return redirect(url_for('view_class', class_name=current_user.assigned_class))
    
    db.session.delete(student)
    db.session.commit()
    flash('Student deleted successfully!', 'success')
    return redirect(url_for('view_class', class_name=current_user.assigned_class))

@app.route('/download_all_report_cards')
@login_required
def download_all_report_cards():
    class_name = current_user.assigned_class
    students = Student.query.filter_by(class_name=class_name).order_by(Student.roll_no).all()
    
    # Create a directory to store individual report cards
    if not os.path.exists('temp_reports'):
        os.makedirs('temp_reports')
    
    # Helper function to safely set cell values for both regular and merged cells
    def set_cell_value(ws, row, col, value):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            # Find the master cell of the merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the master cell (top-left cell of merged range)
                    master_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    master_cell.value = value
                    return
            # If we get here, the cell is merged but we couldn't find its range
            # Try to write to the cell anyway
            try:
                cell.value = value
            except AttributeError:
                print(f"Warning: Could not write to merged cell at row {row}, column {col}")
        else:
            cell.value = value
    
    # Generate report cards for each student using the template
    report_files = []
    try:
        for student in students:
            # Load template for each student
            template_path = os.path.join(basedir, 'static', 'templates', 'report_template.xlsx')
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Fill student details
            set_cell_value(ws, 5, 3, student.name)  # Name in column C
            set_cell_value(ws, 5, 5, student.class_name)  # Class in column E
            set_cell_value(ws, 5, 7, student.roll_no)  # Roll No in column G
            set_cell_value(ws, 5, 9, student.admission_no)  # Admission No in column I
            set_cell_value(ws, 6, 3, student.fathers_name)  # Father's Name in column C
            
            # Subject rows are already defined in template
            subjects = [
                ('ENGLISH', 10, 'english'),
                ('HINDI', 11, 'hindi'),
                ('MATHS', 12, 'maths'),
                ('SCIENCE', 13, 'science'),
                ('SOCIAL SCIENCE', 14, 'sst'),
                ('CO-CURRICULAR', 15, 'cocurricular')  # Added co-curricular subject
            ]
            
            # Components and their columns (B=2, C=3, etc.)
            components = [
                ('pa', 2),      # PA in column B
                ('ma', 3),      # MA in column C
                ('portfolio', 4), # Portfolio in column D
                ('se', 5),      # SE in column E
                ('term', 6)     # Term in column F
            ]
            
            # Fill marks for each subject
            for subject_name, row, code in subjects:
                for component, col in components:
                    mark = getattr(student, f"{code}_{component}", 0) or 0
                    set_cell_value(ws, row, col, mark)
            
            # Fill co-scholastic grades
            set_cell_value(ws, 17, 3, student.gk_grade or 'C')      # GK in column C
            set_cell_value(ws, 17, 6, student.computer_grade or 'C') # Computer in column F
            set_cell_value(ws, 17, 9, student.drawing_grade or 'C')  # Drawing in column I
            
            # Save individual report card
            report_filename = os.path.join('temp_reports', f"report_card_{student.name}_{student.class_name}.xlsx")
            wb.save(report_filename)
            report_files.append(report_filename)
            wb.close()
        
        # Create a zip file containing all report cards
        zip_filename = f"Class_{class_name}_Report_Cards.zip"
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            for report_file in report_files:
                zipf.write(report_file, os.path.basename(report_file))
        
        # Clean up individual report files
        for report_file in report_files:
            try:
                os.remove(report_file)
            except:
                pass
        os.rmdir('temp_reports')
        
        # Return the zip file
        return send_file(
            zip_filename,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
    except Exception as e:
        print(f"Error generating report cards: {str(e)}")
        flash('Error generating report cards. Please try again.', 'error')
        return redirect(url_for('view_class', class_name=class_name))
    finally:
        # Clean up zip file
        try:
            os.remove(zip_filename)
        except:
            pass

def create_default_user():
    with app.app_context():
        if not Teacher.query.filter_by(username='admin').first():
            default_teacher = Teacher(
                username='admin',
                assigned_class='3',
                name='Admin'  
            )
            default_teacher.password = bcrypt.generate_password_hash('admin123').decode('utf-8')
            db.session.add(default_teacher)
            db.session.commit()

if __name__ == '__main__':
    with app.app_context():
        try:
            # Create the instance directory if it doesn't exist
            if not os.path.exists('instance'):
                os.makedirs('instance')
            
            # Create all database tables
            db.create_all()
            print("Database created successfully!")
            
            # Create default admin user
            create_default_user()
            print("Default admin user created!")
            
        except Exception as e:
            print(f"Error setting up database: {str(e)}")
    
    app.run(debug=True)

if __name__ == '__main__':
    with app.app_context():
        try:
            # Create the instance directory if it doesn't exist
            if not os.path.exists('instance'):
                os.makedirs('instance')
            
            # Create all database tables
            db.create_all()
            print("Database created successfully!")
            
            # Create default admin user
            create_default_user()
            print("Default admin user created!")
            
        except Exception as e:
            print(f"Error setting up database: {str(e)}")
    
    app.run(debug=True)
